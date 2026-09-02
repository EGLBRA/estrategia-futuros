# -*- coding: utf-8 -*-
"""v30: itens do validador de fatos v2 (dono do canal, memória de cálculo do KR, selos, figuras dos sintomas regeneradas
com todas as 72 linhas e faixas, núcleo com oito papéis e a facilitação, sups)."""
import io, os, re, datetime
import openpyxl
W = r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01"
X = r"A:\_01 Projetos\Estrategia\Pesquisas\Sintomas Organizacionais GT Susep Vida e RE - 26-09-01\Sintomas_Organizacionais_Susep_Vida_RE.xlsx"
def load(n): return io.open(os.path.join(W, n), encoding="utf-8").read()
def save(n, s): io.open(os.path.join(W, n), "w", encoding="utf-8", newline="\n").write(s)
SV = '<span class="sel sel-v">Verificado</span>'; SI = '<span class="sel sel-i">Inferência</span>'; SP = '<span class="sel sel-p">Proposta</span>'
ST = 'style="display:block;width:100%;height:auto;font-family:system-ui,sans-serif"'
MISS = []
F = {k: load(f"frag_{k}.html") for k in "ABCDEFGHIJKL"}
def R(k, old, new, rx=False, count=0):
    s = F[k]
    if rx: s2, n = re.subn(old, new, s, count=count, flags=re.S)
    else: n = s.count(old); s2 = s.replace(old, new) if n else s
    if n == 0: MISS.append((k, old[:70]))
    F[k] = s2
def RA(old, new, rx=False):
    n = 0
    for k in F:
        if rx: F[k], m = re.subn(old, new, F[k], flags=re.S)
        else: m = F[k].count(old); F[k] = F[k].replace(old, new)
        n += m
    if n == 0: MISS.append(("*", old[:70]))

# 1. dono do canal Unimed
R("L", r"Quem é o dono do canal Unimed e do pós-venda de pessoa física\.[^<]*?<sup>15, 17</sup>[^<]*", "Quem responde pelo pós-venda de pessoa física. A coordenação do canal Unimed tem responsável desde junho; o que falta é a estrutura comercial formalizada, com prazo a definir.<sup>15, 17, 22</sup>", rx=True)
R("A", "Canal sem dono escrito até julho; a proposta nomeia um.", "Coordenação do canal nomeada em junho; estrutura comercial ainda por formalizar.")
# 2. memória de cálculo do KR
RA("ou comunicado a ele.<sup>3, 6, 20, 24</sup>", "ou comunicado a ele; a origem apontada é o planejamento estratégico.<sup>1, 3, 6, 20, 24</sup>")
RA("ou comunicado.<sup>3, 6, 20, 24</sup>", "ou comunicado; a origem apontada é o planejamento estratégico.<sup>1, 3, 6, 20, 24</sup>")
RA("de onde vem o número do KR", "a memória de cálculo do número do KR")
RA("O número do KR não tem origem registrada", "O número do KR não tem memória de cálculo registrada")
RA("A origem do número do KR não está registrada nas atas", "A memória de cálculo do KR não está nas atas")
RA("O número não tem origem registrada.", "O número não tem memória de cálculo.")
RA("Número sem origem explicada", "Número sem memória de cálculo")
RA("nenhuma ata diz a memória de cálculo do número do KR", "nenhuma ata registra a memória de cálculo do número do KR")
RA("as atas não registram a memória de cálculo do número do KR", "as atas não registram a memória de cálculo do número do KR")
# 3. selo da pergunta que separa as camadas
R("I", r"(Uma pergunta separa as camadas[^<]*?)" + re.escape(SV), r"\1" + SP, rx=True, count=1)
# 6. núcleo: oito papéis e a facilitação
R("B", "NÚCLEO: NOVE PAPÉIS, LIDERANÇA ROTATIVA ENTRE LÍDER E COLÍDER", "NÚCLEO: OITO PAPÉIS DE NEGÓCIO E A FACILITAÇÃO; LIDERANÇA ROTATIVA ENTRE LÍDER E COLÍDER")
R("A", '<text x="550" y="262" text-anchor="middle" font-size="12.5" fill="#5a6068">9 papéis, 3 mentores,</text><text x="550" y="279" text-anchor="middle" font-size="12.5" fill="#5a6068">1 facilitação</text>', '<text x="550" y="262" text-anchor="middle" font-size="12.5" fill="#5a6068">9 no núcleo, com a facilitação;</text><text x="550" y="279" text-anchor="middle" font-size="12.5" fill="#5a6068">3 mentores</text>')
# 7, 10, 14, 15 Riscos
R("L", "sem alçada compartilhada e sem via de reconsideração registrada", "sem alçada compartilhada; as atas não registram via de reconsideração")
R("L", "as mesmas pessoas em todos os fóruns", "quase as mesmas pessoas em todos os fóruns")
R("L", "<sup>20, 22, 27</sup>", "<sup>20, 22, 26, 27</sup>")
R("L", "a frente discute cotador e campanha", "a frente só pôde propor a revisão da aceitação em agosto, para deliberação nos fóruns competentes")
# 8. Entenda sup
R("A", r"(nos dois sentidos[^<]*)<sup>27</sup>", r"\1<sup>26, 27</sup>", rx=True, count=1)
# 11. encaminhamento, não decisão
RA("a decisão sai em 36 minutos", "o encaminhamento sai em 36 minutos"); RA("quando a janela existe, a decisão sai", "quando a janela existe, o encaminhamento sai")
# 12, 18, 19
R("A", "receio de mais um modelo dito internamente.<sup>15, 26</sup>", "receio de mais um modelo dito internamente.<sup>26</sup>")
R("I", "concentrados em uma pessoa", "concentrados na facilitação")
R("I", "tem 51 encontros.", "tem 51 encontros, fora dos ritos trimestrais (PIE e abertura das esteiras).")
# 13. Concred
R("B", r",?\s*Concred\b", "", rx=True)
# 17. legenda da distribuição
R("B", "Organização do Trabalho e Informação concentram o volume; Governança e Recursos concentram os que travam decisão. " + SV, "Organização do Trabalho e Informação concentram o volume. " + SV + " Governança e Recursos concentram os que travam decisão. " + SI)

# ---- figuras dos sintomas regeneradas (72 linhas)
MES = {"jan": 1, "fev": 2, "mar": 3, "abr": 4, "mai": 5, "jun": 6, "jul": 7, "ago": 8, "set": 9, "out": 10, "nov": 11, "dez": 12}
def d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    t = str(v or "").strip().lower()
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", t)
    if m: return datetime.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    m = re.match(r"([a-z]{3})[a-z]*/?\s*(\d{4})", t)
    if m and m.group(1) in MES: return datetime.date(int(m.group(2)), MES[m.group(1)], 1)
    return None
wb = openpyxl.load_workbook(X, data_only=True); ws = wb["Sintomas"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ix = {k: hdr.index(k) for k in hdr}
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    rows.append(dict(cat=r[ix["Categoria"]], sint=r[ix["Sintoma / Problema"]], n=int(r[ix["Nº de fontes"]] or 0), a=d(r[ix["Primeira menção"]]), b=d(r[ix["Última menção"]])))
cats = ["Organização do Trabalho", "Governança", "Coordenação", "Informação", "Recursos", "Tecnologia"]
CC = {"Organização do Trabalho": "#1a4a8a", "Governança": "#b71c1c", "Coordenação": "#b47c00", "Informação": "#6410ab", "Recursos": "#00995d", "Tecnologia": "#5b626a"}
def fig(svg, cap): return '<figure style="margin:14px 0 22px">\n' + svg + '\n<figcaption class="note" style="margin-top:6px">' + cap + '</figcaption>\n</figure>\n'
def esc(t): return str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
# temperatura
months = [(2026, m) for m in range(3, 9)]; mname = ["mar", "abr", "mai", "jun", "jul", "ago"]
grid = {c: [0] * 6 for c in cats}; n = 0
for r in rows:
    if r["cat"] not in grid or not r["a"]: continue
    b = r["b"] or r["a"]; n += 1
    for k, (y, m) in enumerate(months):
        first = datetime.date(y, m, 1); last = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
        if r["a"] <= last and b >= first: grid[r["cat"]][k] += 1
mx = max(max(v) for v in grid.values())
def col(v):
    if v == 0: return "#f4f5f7"
    t = v / mx; return "#%02x%02x%02x" % (255, int(236 - 170 * t), int(230 - 190 * t))
s = '<svg viewBox="0 0 1100 330" role="img" aria-label="Temperatura dos sintomas: quantos sintomas de cada categoria estavam ativos em cada mês, de março a agosto de 2026" xmlns="http://www.w3.org/2000/svg" %s>' % ST
s += '<text x="20" y="28" font-size="15" font-weight="700" fill="#1a1a1a">Temperatura dos sintomas, março a agosto de 2026</text><text x="20" y="48" font-size="12.5" fill="#5a6068">cada célula conta os sintomas da categoria ativos no mês (entre a primeira e a última menção nas atas); quanto mais escuro, mais quente</text>'
x0, y0, cw, ch = 300, 70, 120, 36
for k, mn in enumerate(mname): s += '<text x="%d" y="%d" text-anchor="middle" font-size="13" font-weight="600" fill="#5a6068">%s</text>' % (x0 + k * cw + cw / 2, y0 - 8, mn)
tot = [0] * 6
for i, c in enumerate(cats):
    y = y0 + i * ch; s += '<text x="20" y="%d" font-size="13.5" fill="#1a1a1a">%s</text>' % (y + 23, c)
    for k, v in enumerate(grid[c]):
        tot[k] += v
        s += '<rect x="%d" y="%d" width="%d" height="%d" fill="%s" stroke="#fff"/><text x="%d" y="%d" text-anchor="middle" font-size="13.5" font-weight="600" fill="%s">%d</text>' % (x0 + k * cw, y, cw, ch, col(v), x0 + k * cw + cw / 2, y + 23, "#1a1a1a" if v < mx * 0.6 else "#fff", v)
y = y0 + 6 * ch + 6
s += '<line x1="20" y1="%d" x2="1040" y2="%d" stroke="#d7dbe0"/><text x="20" y="%d" font-size="13.5" font-weight="700" fill="#1a1a1a">Sintomas ativos no mês</text>' % (y, y, y + 24)
for k, v in enumerate(tot): s += '<text x="%d" y="%d" text-anchor="middle" font-size="14" font-weight="700" fill="#b71c1c">%d</text>' % (x0 + k * cw + cw / 2, y + 24, v)
s += '</svg>'
FIG_T = fig(s, "Contagem deste documento a partir das datas de primeira e última menção dos %d sintomas na planilha (o que só tem mês de referência entra pelo primeiro dia do mês).<sup>25</sup> A temperatura sobe até junho e segue alta em julho e agosto. " % n + SI)
# matriz por faixas
def band_p(v): return 0 if v < 1.5 else 1 if v < 2.5 else 2 if v < 4.5 else 3
def band_a(v): return 0 if v < 1.5 else 1 if v < 2.5 else 2 if v < 4.5 else 3
agg = {}
for r in rows:
    if r["cat"] not in cats or not r["a"]: continue
    b = r["b"] or r["a"]; mo = (b.year - r["a"].year) * 12 + b.month - r["a"].month + 1
    a = agg.setdefault(r["cat"], [0, 0, 0]); a[0] += mo; a[1] += r["n"]; a[2] += 1
s = '<svg viewBox="0 0 1100 460" role="img" aria-label="Matriz com as seis categorias posicionadas por faixa de persistência média em meses e de abrangência média em documentos; o tamanho da bolha é o número de sintomas" xmlns="http://www.w3.org/2000/svg" %s>' % ST
gx0, gy0, gw, gh = 140, 50, 900, 340; cols4 = ["#e8f5e9", "#fff9c4", "#ffe0b2", "#ffcdd2"]
for i in range(4):
    for j in range(4):
        s += '<rect x="%d" y="%d" width="%d" height="%d" fill="%s" stroke="#fff" stroke-width="3"/>' % (gx0 + j * gw / 4, gy0 + (3 - i) * gh / 4, gw / 4, gh / 4, cols4[min(3, (i + j) // 2)])
for j, lab in enumerate(["1 documento", "2", "3 a 4", "5 ou mais"]): s += '<text x="%d" y="%d" text-anchor="middle" font-size="12.5" fill="#5a6068">%s</text>' % (gx0 + j * gw / 4 + gw / 8, gy0 + gh + 22, lab)
for i, lab in enumerate(["1 mês", "2", "3 a 4", "5 a 6"]): s += '<text x="%d" y="%d" text-anchor="end" font-size="12.5" fill="#5a6068">%s</text>' % (gx0 - 12, gy0 + (3 - i) * gh / 4 + gh / 8 + 4, lab)
s += '<text x="%d" y="%d" text-anchor="middle" font-size="13" font-weight="600" fill="#1a1a1a">Abrangência: documentos que citam o sintoma (média da categoria)</text>' % (gx0 + gw / 2, gy0 + gh + 44)
s += '<text x="30" y="%d" font-size="13" font-weight="600" fill="#1a1a1a" transform="rotate(-90 30 %d)" text-anchor="middle">Persistência: meses ativo (média da categoria)</text>' % (gy0 + gh / 2, gy0 + gh / 2)
placed = {}
for c in cats:
    if c not in agg: continue
    pm, nm, k = agg[c][0] / agg[c][2], agg[c][1] / agg[c][2], agg[c][2]
    bi, bj = band_p(pm), band_a(nm); key = (bi, bj); off = placed.get(key, 0); placed[key] = off + 1
    x = gx0 + bj * gw / 4 + gw / 8 + (off % 2) * 70 - (35 if off else 0); y = gy0 + (3 - bi) * gh / 4 + gh / 8 - 10 + (off // 2) * 40
    rr = 14 + k * 1.6
    s += '<circle cx="%d" cy="%d" r="%d" fill="%s" opacity="0.88"/><text x="%d" y="%d" text-anchor="middle" font-size="12.5" font-weight="700" fill="#fff">%d</text><text x="%d" y="%d" text-anchor="middle" font-size="11.5" fill="#1a1a1a">%s (%.1f m · %.1f doc)</text>' % (x, y, rr, CC[c], x, y + 4, k, x, y + rr + 14, c.replace("Organização do Trabalho", "Org. do Trabalho"), pm, nm)
s += '</svg>'
FIG_M = fig(s, "Cada bolha é uma categoria, posicionada na faixa da sua média; o número é quantos sintomas ela tem (as seis somam %d). Quanto mais alto e à direita, mais tempo o problema fica vivo e em mais documentos ele aparece. Médias calculadas da planilha.<sup>25</sup> " % sum(v[2] for v in agg.values()) + SI)
# barras: todos com 4 ou mais documentos
top = sorted([r for r in rows if r["n"] >= 4], key=lambda r: (-r["n"], str(r["sint"])))
mxn = max(r["n"] for r in top); H = 60 + len(top) * 28 + 70
s = '<svg viewBox="0 0 1100 %d" role="img" aria-label="Os problemas citados em quatro ou mais documentos, com a cor indicando a dimensão organizacional" xmlns="http://www.w3.org/2000/svg" %s>' % (H, ST)
s += '<text x="20" y="28" font-size="15" font-weight="700" fill="#1a1a1a">Problema e quantidade de documentos em que aparece</text><text x="20" y="48" font-size="12.5" fill="#5a6068">os %d problemas citados em quatro ou mais documentos, entre os 72; a cor é o recorte organizacional</text>' % len(top)
x0b, bw = 470, 560
for i, r in enumerate(top):
    y = 64 + i * 28; nm = str(r["sint"]); lab = esc(nm if len(nm) <= 62 else nm[:59].rsplit(" ", 1)[0])
    w = bw * r["n"] / mxn
    s += '<text x="%d" y="%d" text-anchor="end" font-size="12.5" fill="#1a1a1a">%s</text><rect x="%d" y="%d" width="%d" height="19" rx="3" fill="%s"/><text x="%d" y="%d" font-size="12.5" font-weight="700" fill="#1a1a1a">%d</text>' % (x0b - 10, y + 15, lab, x0b, y + 2, w, CC.get(r["cat"], "#999"), x0b + w + 8, y + 16, r["n"])
ly = 64 + len(top) * 28 + 22; lx = 20
for cat, c in CC.items():
    s += '<rect x="%d" y="%d" width="14" height="14" rx="3" fill="%s"/><text x="%d" y="%d" font-size="12" fill="#1a1a1a">%s</text>' % (lx, ly, c, lx + 20, ly + 12, cat); lx += 24 + 8 * len(cat) + 20
s += '</svg>'
FIG_B = fig(s, "Contagem de documentos que citam cada problema, na planilha de sintomas; empates em ordem alfabética.<sup>25</sup> Os mais citados se espalham pelas seis dimensões: o padrão é do desenho, não de uma área. " + SV)
B = F["B"]
for label, new in [("Temperatura dos sintomas", FIG_T), ("Matriz com as seis categorias", FIG_M), ("Os dezesseis problemas", FIG_B)]:
    m = re.search(r'<figure style="margin:14px 0 22px">\s*<svg viewBox[^>]*aria-label="%s.*?</figure>\n' % re.escape(label), B, re.S)
    if m: B = B.replace(m.group(0), new, 1)
    else: MISS.append(("B", "figura " + label))
F["B"] = B
for k, s_ in F.items(): save(f"frag_{k}.html", s_)
print("patch v30 ok; faltas:", MISS, "| temperatura n:", n, "| barras:", len(top))

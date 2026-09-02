# -*- coding: utf-8 -*-
"""v6: abas (remove STEEPLEDI, Porter, Check; adiciona Visão executiva, Proposta, Reuniões e custo),
'Na prática' em cada aba, tarjeta EXEMPLO, selo Proposta nas abas de proposta, custo das reuniões
(HTML + Excel com parâmetro), limites de método em Fontes."""
import io, os, re, sys
sys.path.insert(0, r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01")
from pratica import PRATICA
W = r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01"
OUTDIR = r"A:\_01 Projetos\Estrategia\Pesquisas\Anatomia Profunda - GT Susep Vida e RE - 26-09-01"
def load(n): return io.open(os.path.join(W, n), encoding="utf-8").read()
def save(n, s): io.open(os.path.join(W, n), "w", encoding="utf-8", newline="\n").write(s)
def rep(s, a, b, must=False):
    if a not in s:
        if must: raise AssertionError("nao achou: " + a[:90])
        return s
    return s.replace(a, b, 1)

# ---------------- custo das reuniões (valor-hora parametrizado)
VH = 125
PROP = [  # rito, sessoes/ano, duracao h, pessoas
    ("Planejamento estratégico", 1, 8, 12), ("PIE", 4, 4, 15), ("Fórum de Negócio SUSEP", 12, 2, 15),
    ("Mentoria por tema", 12, 1, 6), ("Abertura da esteira (Vida e RE)", 8, 2, 8), ("Triagem de demanda", 24, 0.5, 8),
    ("Comitê de priorização", 12, 1, 10), ("Reunião da esteira (Vida e RE)", 96, 1.5, 8),
    ("Sala de descoberta (6 desafios por ano)", 36, 2, 4), ("Revisão do time dedicado", 24, 0.75, 8),
    ("Reunião diária do núcleo", 220, 0.25, 3), ("Sala de guerra (4 temas por ano)", 32, 1, 6),
]
HOJE = [
    ("Reunião semanal mista Vida e RE", 48, 1.5, 10), ("S&amp;OP Vida", 12, 2, 10), ("Diagnóstico SUSEP", 12, 2, 10),
    ("Fórum de Gestão Estratégica (até 4h)", 12, 4, 15), ("Mentoria (30 min)", 12, 0.5, 6),
    ("PIE", 4, 4, 15), ("Planejamento estratégico", 1, 8, 12),
]
def fmt(n): return f"{n:,.0f}".replace(",", ".")
def tabela(rows, titulo):
    h = [f"<h4>{titulo}</h4>", "<table><thead><tr><th>Rito</th><th>Sessões no ano</th><th>Duração (h)</th><th>Pessoas</th><th>Horas-pessoa no ano</th><th>Custo no ano a R$ 125/h</th></tr></thead><tbody>"]
    tot_h = tot_c = 0
    for r, s, d, p in rows:
        hp = s * d * p; c = hp * VH; tot_h += hp; tot_c += c
        h.append(f"<tr><td>{r}</td><td>{s}</td><td>{d:g}</td><td>{p}</td><td>{fmt(hp)}</td><td>R$ {fmt(c)}</td></tr>")
    h.append(f"<tr><td><strong>Total</strong></td><td></td><td></td><td></td><td><strong>{fmt(tot_h)}</strong></td><td><strong>R$ {fmt(tot_c)}</strong></td></tr></tbody></table>")
    return "\n".join(h), tot_h, tot_c
t1, h1, c1 = tabela(PROP, "Sistema proposto")
t2, h2, c2 = tabela(HOJE, "Sistema atual (estimado a partir das atas)")
CUSTO = f'<div class="card"><strong>Parâmetro: valor-hora médio da sala = R$ {VH}</strong> (padrão; a planilha anexa recalcula com outro valor).</div>\n{t1}\n{t2}\n<p><strong>Diferença:</strong> o sistema proposto custa cerca de R$ {fmt(c1 - c2)} a mais por ano ({fmt(h1 - h2)} horas-pessoa), a R$ {VH} a hora. O que compra: triagem, descoberta, revisão de resultado e esteiras separadas. O que elimina: três agendas paralelas e o fórum de quatro horas.</p>'
print("custo proposto", c1, "atual", c2)

# ---------------- fragmentos
G = load("frag_G.html"); H = load("frag_H.html")
G = rep(G, "<!-- CUSTO_TABELA -->", '<p>A tabela completa, com o sistema atual para comparação, está na aba Reuniões e custo. Resumo: sistema proposto perto de R$ ' + fmt(c1) + ' por ano; sistema atual estimado perto de R$ ' + fmt(c2) + '; parâmetro R$ 125 a hora.</p>')
G = rep(G, "custa cerca de R$ 380 mil por ano a R$ 125 a hora, contra cerca de R$ 300 mil do sistema atual estimado", f"custa cerca de R$ {fmt(round(c1, -3))} por ano a R$ 125 a hora, contra cerca de R$ {fmt(round(c2, -3))} do sistema atual estimado")
H = rep(H, "<!-- CUSTO_TABELA -->", CUSTO)
# Síntese vira Visão executiva
G = rep(G, '<section id="p-sintese" class="pane">\n<h2>A síntese: o diagnóstico em uma página</h2>', '<section id="p-exec" class="pane">\n<h2>Visão executiva: o diagnóstico e a proposta em uma página</h2>')
G = rep(G, '<h3>Por onde começar</h3>', '<h3>A proposta, em duas linhas</h3>\n<p>Construir o desenho com o próprio grupo em seis sessões, em seis semanas, partindo dos problemas que ele registrou: fluxo da estratégia à entrega, fluxo funcional entre áreas, sistema de reuniões com custo, alçadas e regras, e só então papéis. Piloto em Vida por 90 dias, medido (aba Proposta).</p>\n<h3>Por onde começar</h3>')
G = rep(G, "<!-- ============================== SINTESE ============================== -->", "<!-- ============================== EXEC ============================== -->")
save("frag_G.html", G); save("frag_H.html", H)

# frag_E: substituir a secao linear pela nova (frag_G) e manter fluxo funcional; tarjeta EXEMPLO
E = load("frag_E.html")
E = re.sub(r'<!-- =+ LINEAR =+ -->\s*<section id="p-linear" class="pane">.*?</section>\s*', '', E, count=1, flags=re.S)

E = rep(E, '<div style="background:#d0342c;color:#fff;padding:16px 20px;border-radius:6px;margin:14px 0;display:flex;gap:20px;align-items:center;flex-wrap:wrap">\n<b style="font-size:30px;letter-spacing:6px;font-weight:800;line-height:1;font-family:system-ui,sans-serif">RASCUNHO</b>\n<span style="font-size:14px;max-width:900px;line-height:1.45">Nada aqui está decidido. Parte do esboço da Estratégia (Planejamento, Mapa, Squads, Fórum de Negócio, Aceleradores, Projetos de torre) e acrescenta o que faltava nele: de onde nasce o problema, quando acontece o discovery, para onde a solução vai e onde entram RDS e COMEX. Versão 0.2.</span>\n</div>',
    '<div class="alerta"><strong>EXEMPLO PARA DISCUSSÃO.</strong> Parte do esboço da Estratégia (Planejamento, Mapa, times dedicados, Fórum de Negócio, Aceleradores, Projetos de torre) e acrescenta o que faltava nele: de onde nasce o problema, quando acontece a descoberta, para onde a solução vai e onde entram RDS e COMEX. O desenho válido é o que sair das sessões com o time (aba Proposta).</div>', must=False)
E = re.sub(r'<text x="700" y="440" text-anchor="middle" font-size="150"[^>]*>RASCUNHO</text>', '<text x="700" y="440" text-anchor="middle" font-size="110" font-weight="800" fill="#1a1a1a" fill-opacity="0.05" transform="rotate(-16 700 440)" letter-spacing="16">EXEMPLO</text>', E)
E = E.replace("Nada aqui está decidido. ", "")
E = rep(E, "<h2>Fluxo funcional: quem alimenta quem, da origem do problema à solução entregue</h2>", '<h2>Fluxo funcional: quem alimenta quem, da origem do problema à solução entregue <span class="sel sel-p">Exemplo</span></h2>')
save("frag_E.html", E)
G = load("frag_G.html")
G = rep(G, "<h2>Da estratégia à entrega: o caminho de um desafio, etapa por etapa</h2>", '<h2>Da estratégia à entrega: o caminho de um desafio, etapa por etapa <span class="sel sel-p">Exemplo</span></h2>')
G = rep(G, '<div class="alerta">Proposta para discussão. Cadências, prazos, limites e o custo das salas são pontos de partida para o time calibrar. <span class="sel sel-e">Especulativo</span> nos números; <span class="sel sel-i">Inferência</span> no desenho, derivado dos 72 sintomas.<sup>25</sup></div>',
    '<div class="alerta"><strong>EXEMPLO PARA DISCUSSÃO.</strong> Cadências, prazos, limites e o custo das salas são pontos de partida para o time calibrar nas sessões (aba Proposta). O desenho deriva dos 72 sintomas.<sup>25</sup></div>')
save("frag_G.html", G)

# ---------------- selo Proposta nas abas de proposta (sem inferência)
PROPOSAL_TABS = {"E": None, "G": None, "H": None, "C": ("p-modelo", "p-blue", "p-5w2h"), "F": ("p-design",)}
def to_proposta(html):
    html = html.replace('<span class="sel sel-i">Inferência</span>', '<span class="sel sel-p">Proposta</span>')
    html = html.replace('<span class="sel sel-e">Especulativo</span>', '<span class="sel sel-p">Proposta</span>')
    html = re.sub(r'\s*<span class="sel sel-p">Proposta</span>(\s+(?:sobre|na|no|nos|nas|quanto|a|o|em)\b[^.<]*)', r' <span class="sel sel-p">Proposta</span>', html)
    html = re.sub(r'(<span class="sel sel-p">Proposta</span>)(\s*<span class="sel sel-p">Proposta</span>)+', r'', html)
    html = html.replace('<span class="sel sel-p">Proposta</span><span class="sel sel-v">Verificado</span> nos prazos', '<span class="sel sel-p">Proposta</span> no custo; <span class="sel sel-v">Verificado</span> nos prazos')
    html = re.sub(r'<p>\s*<span class="sel sel-p">Proposta</span>\.?\s*</p>', '<p><span class="sel sel-p">Proposta</span></p>', html)
    html = re.sub(r'<span class="sel sel-p">Proposta</span>\s*\.', '<span class="sel sel-p">Proposta</span>', html)
    return html
for k, secs in PROPOSAL_TABS.items():
    s = load(f"frag_{k}.html")
    if secs is None:
        s = to_proposta(s)
    else:
        for sec in secs:
            m = re.search(r'(<section id="%s" class="pane">.*?</section>)' % sec, s, re.S)
            s = s.replace(m.group(1), to_proposta(m.group(1)))
    save(f"frag_{k}.html", s)

# ---------------- Na prática em cada aba
frags = {k: load(f"frag_{k}.html") for k in "ABCDEFGH"}
def insert_pratica(html, pid, box):
    pat = re.compile(r'(<section id="p-%s" class="pane">.*?)(</section>)' % pid, re.S)
    m = pat.search(html); assert m, pid
    if 'class="pratica"' in m.group(1): return html
    return html[:m.start()] + m.group(1) + "\n" + box + "\n" + m.group(2) + html[m.end():]
for pid, box in PRATICA.items():
    done = False
    for k in frags:
        if f'id="p-{pid}"' in frags[k]:
            frags[k] = insert_pratica(frags[k], pid, box); done = True; break
    assert done, pid
for k, s in frags.items(): save(f"frag_{k}.html", s)

# ---------------- Fontes: limites de método (o que o Check dizia e não se resolve com texto)
D = load("frag_D.html")
D = rep(D, "Nenhum número, citação ou data foi criado fora delas.",
        "Nenhum número, citação ou data foi criado fora delas. Limites que o leitor deve carregar: as fontes são internas e produzidas pelo próprio grupo, sem medição de sistema; \"115% e 99%\" vêm de uma ata de mentores sem a base numérica ao lado; \"36% dos corretores\" é um recorte de São Paulo; \"5 anos de defasagem\" é estimativa do próprio grupo; \"13\" do mentor não tem unidade registrada; a cifra da Icatu é citação de terceiro em ata; a contagem de 72 sintomas separa itens correlatos (agenda, ritos com mentores e engajamento entre reuniões são faces do mesmo tempo escasso); o elemento vigilância não tem evidência; a ata de 4 de maio está duplicada (fontes 7 e 8) e a de 27 e 28 de julho traz um bloco repetido (fonte 24); as duas transcrições são automáticas, citadas por função. O autor esteve na reunião da fonte 26 e propôs a leitura que o documento defende; por isso só as falas das participantes da Estratégia são usadas como evidência.")
save("frag_D.html", D)

# ---------------- build.py: abas, fragmentos, CSS do selo Proposta e do h4
B = load("build.py")
B = re.sub(r'TABS = \[.*?\]\n', '''TABS = [("dest", "Destaque"), ("exec", "Visão executiva"), ("proposta", "Proposta"), ("entenda", "Entenda"), ("eco", "Ecossistema"), ("proc", "Processos"),
        ("sint", "Sintomas"), ("normas", "Normas e regras"), ("papeis", "Papéis e pessoas"), ("sist", "Sistemas e dados"),
        ("loops", "Loops"), ("modelo", "Modelo"), ("design", "Perguntas de design"), ("blue", "Blueprint"),
        ("linear", "Da estratégia à entrega"), ("reun", "Fluxo funcional"), ("ritos", "Reuniões e custo"), ("5w2h", "5W2H"),
        ("analise", "Análise"), ("sut", "Sutilezas"), ("cem", "Cem perguntas"), ("gloss", "Glossário"), ("fontes", "Fontes")]
''', B, count=1, flags=re.S)
B = rep(B, 'for x in "ABCDEF")', 'for x in "ABCDEFGH")')
B = rep(B, "grid-template-columns:repeat(12,max-content)", "grid-template-columns:repeat(12,max-content)", must=False)
if ".sel-p{" not in B:
    B = rep(B, 'head = head.replace("grid-template-columns:repeat(11,max-content)"', 'head = head.replace("</style>", ".sel-p{background:#e8f0fe;color:#1a4a8a}\\nh4{font-size:14px;font-weight:700;margin:14px 0 4px}\\n</style>")\nhead = head.replace("grid-template-columns:repeat(11,max-content)"')
B = rep(B, "Este documento reúne o diagnóstico, o blueprint da estratégia à entrega e o rascunho do fluxo funcional entre áreas, grupos e fóruns.",
        "Este documento reúne o diagnóstico, a visão executiva, a proposta de construir o desenho com o time e três exemplos de partida: o fluxo da estratégia à entrega, o fluxo funcional entre áreas e o sistema de reuniões com custo.")
save("build.py", B)

# ---------------- Excel do custo com parâmetro
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
wb = Workbook(); ws = wb.active; ws.title = "Custo"
thin = Side(style="thin", color="BFBFBF"); border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr = Font(name="Arial", bold=True, color="FFFFFF"); fill = PatternFill("solid", fgColor="1F3864"); body = Font(name="Arial", size=10)
ws["A1"] = "CUSTO DO SISTEMA DE REUNIÕES · SUSEP Vida e RE · proposta x atual"; ws["A1"].font = Font(name="Arial", bold=True, size=12)
ws["A3"] = "Valor-hora médio da sala (R$)"; ws["A3"].font = Font(name="Arial", bold=True)
ws["B3"] = VH; ws["B3"].font = Font(name="Arial", color="0000FF", bold=True); ws["B3"].fill = PatternFill("solid", fgColor="FFFF00"); ws["B3"].number_format = "#,##0"
ws["C3"] = "Célula de entrada: altere e tudo recalcula"; ws["C3"].font = Font(name="Arial", italic=True, size=9)
row = 5
def bloco(titulo, rows):
    global row
    ws.cell(row=row, column=1, value=titulo).font = Font(name="Arial", bold=True); row += 1
    for i, h in enumerate(["Rito", "Sessões no ano", "Duração (h)", "Pessoas", "Horas-pessoa no ano", "Custo no ano (R$)"], start=1):
        c = ws.cell(row=row, column=i, value=h); c.font = hdr; c.fill = fill; c.border = border; c.alignment = Alignment(wrap_text=True)
    row += 1; first = row
    for r, s, d, p in rows:
        ws.cell(row=row, column=1, value=r.replace("&amp;", "&")); ws.cell(row=row, column=2, value=s); ws.cell(row=row, column=3, value=d); ws.cell(row=row, column=4, value=p)
        ws.cell(row=row, column=5, value=f"=B{row}*C{row}*D{row}"); ws.cell(row=row, column=6, value=f"=E{row}*$B$3")
        for col in range(1, 7):
            c = ws.cell(row=row, column=col); c.font = Font(name="Arial", size=10, color="0000FF" if col in (2, 3, 4) else "000000"); c.border = border
        ws.cell(row=row, column=6).number_format = "#,##0"; ws.cell(row=row, column=5).number_format = "#,##0"
        row += 1
    last = row - 1
    ws.cell(row=row, column=1, value="Total").font = Font(name="Arial", bold=True)
    ws.cell(row=row, column=5, value=f"=SUM(E{first}:E{last})").number_format = "#,##0"
    ws.cell(row=row, column=6, value=f"=SUM(F{first}:F{last})").number_format = "#,##0"
    for col in (5, 6): ws.cell(row=row, column=col).font = Font(name="Arial", bold=True)
    tot = row; row += 2
    return tot
t_prop = bloco("Sistema proposto", PROP); t_hoje = bloco("Sistema atual (estimado a partir das atas)", HOJE)
ws.cell(row=row, column=1, value="Diferença anual (proposto menos atual)").font = Font(name="Arial", bold=True)
ws.cell(row=row, column=5, value=f"=E{t_prop}-E{t_hoje}").number_format = "#,##0"; ws.cell(row=row, column=6, value=f"=F{t_prop}-F{t_hoje}").number_format = "#,##0"
row += 2
ws.cell(row=row, column=1, value="Legenda: azul = entrada (sessões, duração, pessoas, valor-hora); preto = fórmula. Pessoas e sessões são hipóteses de partida para o time corrigir. Fonte: Anatomia Profunda GT Susep Vida e RE, 01/09/2026.").font = Font(name="Arial", size=9, italic=True)
for col, w in zip("ABCDEF", [44, 14, 12, 10, 20, 20]): ws.column_dimensions[col].width = w
xl = os.path.join(OUTDIR, "Custo das Reunioes SUSEP - parametro 125.xlsx"); wb.save(xl); print("xlsx", xl)
print("patch v6 ok")

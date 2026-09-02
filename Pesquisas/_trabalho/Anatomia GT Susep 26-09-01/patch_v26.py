# -*- coding: utf-8 -*-
"""v26: na aba Sintomas, a cobertura em estilo de noticiário: distribuição por categoria (barras), matriz persistência x abrangência
por categoria (bolhas) e os sintomas mais recorrentes um a um, em cartões com selo de natureza, números e fala da ata."""
import io, os, re, datetime
import openpyxl
W = r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01"
X = r"A:\_01 Projetos\Estrategia\Pesquisas\Sintomas Organizacionais GT Susep Vida e RE - 26-09-01\Sintomas_Organizacionais_Susep_Vida_RE.xlsx"
def load(n): return io.open(os.path.join(W, n), encoding="utf-8").read()
def save(n, s): io.open(os.path.join(W, n), "w", encoding="utf-8", newline="\n").write(s)
SI = '<span class="sel sel-i">Inferência</span>'; SV = '<span class="sel sel-v">Verificado</span>'
ST = 'style="display:block;width:100%;height:auto;font-family:system-ui,sans-serif"'
def esc(t): return str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
def d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    try: return datetime.datetime.strptime(str(v)[:10], "%d/%m/%Y").date()
    except Exception: return None
wb = openpyxl.load_workbook(X, data_only=True); ws = wb["Sintomas"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ix = {k: hdr.index(k) for k in ["ID", "Categoria", "Elemento", "Sintoma / Problema", "Como este elemento se manifesta na prática", "Evidência nas fontes (citação)", "Fontes", "Nº de fontes", "Primeira menção", "Última menção", "Ramo / Canal afetado", "Tipo de evidência"] if k in hdr}
def col(k, r):
    for h, i in ix.items():
        if h.startswith(k): return r[i]
    return None
rows = []
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0]: continue
    rows.append(dict(id=r[0], cat=col("Categoria", r), sint=col("Sintoma", r), como=col("Como", r), cit=col("Evidência nas fontes", r), fontes=str(col("Fontes", r) or ""), n=int(col("Nº de fontes", r) or 0), a=d(col("Primeira", r)), b=d(col("Última", r)), ramo=col("Ramo", r) or "", tipo=col("Tipo", r) or ""))
cats = ["Organização do Trabalho", "Governança", "Coordenação", "Informação", "Recursos", "Tecnologia"]
CC = {"Organização do Trabalho": "#1a4a8a", "Governança": "#b71c1c", "Coordenação": "#b47c00", "Informação": "#6410ab", "Recursos": "#00995d", "Tecnologia": "#5b626a"}

# ---- (a) barras: sintomas por categoria, com quantos são recorrentes (3 ou mais fontes)
cnt = {c: [0, 0] for c in cats}
for r in rows:
    if r["cat"] in cnt: cnt[r["cat"]][0] += 1; cnt[r["cat"]][1] += (1 if r["n"] >= 3 else 0)
mx = max(v[0] for v in cnt.values())
s = '<svg viewBox="0 0 1100 330" role="img" aria-label="Distribuição dos 72 sintomas por categoria; a parte escura de cada barra são os sintomas citados em três ou mais documentos" xmlns="http://www.w3.org/2000/svg" %s>' % ST
s += '<text x="20" y="28" font-size="15" font-weight="700" fill="#1a1a1a">Distribuição: sintomas por categoria</text><text x="1080" y="28" text-anchor="end" font-size="12.5" fill="#5a6068">n=%d · escuro: citados em 3 ou mais documentos</text>' % len(rows)
bx0, by0, bw, bh = 80, 60, 150, 200
for i, c in enumerate(cats):
    t, rec = cnt[c]; x = bx0 + i * 168; h = bh * t / mx; hr = bh * rec / mx
    s += '<rect x="%d" y="%d" width="%d" height="%d" rx="4" fill="%s" opacity="0.35"/>' % (x, by0 + bh - h, bw, h, CC[c])
    s += '<rect x="%d" y="%d" width="%d" height="%d" rx="4" fill="%s"/>' % (x, by0 + bh - hr, bw, hr, CC[c])
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="14" font-weight="700" fill="#1a1a1a">%d</text>' % (x + bw / 2, by0 + bh - h - 8, t)
    lab = c if len(c) < 14 else c.replace("Organização do ", "Org. do ")
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="12.5" fill="#1a1a1a">%s</text><text x="%d" y="%d" text-anchor="middle" font-size="11.5" fill="#5a6068">%d recorrentes</text>' % (x + bw / 2, by0 + bh + 20, lab, x + bw / 2, by0 + bh + 37, rec)
s += '<line x1="60" y1="%d" x2="1080" y2="%d" stroke="#1a1a1a"/>' % (by0 + bh, by0 + bh)
s += '</svg>'
FIG_A = '<figure style="margin:14px 0 22px">\n' + s + '\n<figcaption class="note" style="margin-top:6px">Contagem da planilha de sintomas.<sup>25</sup> Organização do Trabalho e Informação concentram o volume; Governança e Recursos concentram os que travam decisão. ' + SV + '</figcaption>\n</figure>\n'

# ---- (b) matriz: persistência (meses entre primeira e última menção) x abrangência (documentos que citam), por categoria
agg = {}
for r in rows:
    if r["cat"] not in cats or not r["a"]: continue
    b = r["b"] or r["a"]; months = (b.year - r["a"].year) * 12 + b.month - r["a"].month + 1
    a = agg.setdefault(r["cat"], [0, 0, 0]); a[0] += months; a[1] += r["n"]; a[2] += 1
s = '<svg viewBox="0 0 1100 460" role="img" aria-label="Matriz com as seis categorias posicionadas por persistência média em meses e abrangência média em documentos; o tamanho da bolha é o número de sintomas" xmlns="http://www.w3.org/2000/svg" %s>' % ST
gx0, gy0, gw, gh = 140, 50, 900, 340
cols4 = ["#e8f5e9", "#fff9c4", "#ffe0b2", "#ffcdd2"]
for i in range(4):
    for j in range(4):
        risk = i + j
        s += '<rect x="%d" y="%d" width="%d" height="%d" fill="%s" stroke="#fff" stroke-width="3"/>' % (gx0 + j * gw / 4, gy0 + (3 - i) * gh / 4, gw / 4, gh / 4, cols4[min(3, risk // 2)])
for j, lab in enumerate(["1 documento", "2", "3 a 4", "5 ou mais"]):
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="12.5" fill="#5a6068">%s</text>' % (gx0 + j * gw / 4 + gw / 8, gy0 + gh + 22, lab)
for i, lab in enumerate(["1 mês", "2", "3 a 4", "5 a 6"]):
    s += '<text x="%d" y="%d" text-anchor="end" font-size="12.5" fill="#5a6068">%s</text>' % (gx0 - 12, gy0 + (3 - i) * gh / 4 + gh / 8 + 4, lab)
s += '<text x="%d" y="%d" text-anchor="middle" font-size="13" font-weight="600" fill="#1a1a1a">Abrangência: documentos que citam o sintoma (média)</text>' % (gx0 + gw / 2, gy0 + gh + 44)
s += '<text x="30" y="%d" font-size="13" font-weight="600" fill="#1a1a1a" transform="rotate(-90 30 %d)" text-anchor="middle">Persistência: meses ativo (média)</text>' % (gy0 + gh / 2, gy0 + gh / 2)
def px(v):  # abrangência média 1..6 -> x
    v = max(1, min(6, v)); return gx0 + (v - 1) / 5 * gw * 0.92 + gw * 0.04
def py(v):  # persistência média 1..6 -> y
    v = max(1, min(6, v)); return gy0 + gh - ((v - 1) / 5 * gh * 0.92 + gh * 0.04)
for c in cats:
    if c not in agg: continue
    pm, nm, k = agg[c][0] / agg[c][2], agg[c][1] / agg[c][2], agg[c][2]
    x, y = px(nm), py(pm); rr = 14 + k * 1.6
    s += '<circle cx="%d" cy="%d" r="%d" fill="%s" opacity="0.85"/><text x="%d" y="%d" text-anchor="middle" font-size="12.5" font-weight="700" fill="#fff">%d</text>' % (x, y, rr, CC[c], x, y + 4, k)
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="12" fill="#1a1a1a">%s</text>' % (x, y + rr + 15, c.replace("Organização do Trabalho", "Org. do Trabalho"))
s += '</svg>'
FIG_B = '<figure style="margin:14px 0 22px">\n' + s + '\n<figcaption class="note" style="margin-top:6px">Cada bolha é uma categoria; o número é quantos sintomas ela tem. Quanto mais alto e à direita, mais tempo o problema fica vivo e em mais documentos ele aparece. Médias calculadas da planilha.<sup>25</sup> ' + SI + '</figcaption>\n</figure>\n'

# ---- (c) os sintomas que mais aparecem, um a um
def badge(r):
    ativo_ago = r["b"] and r["b"].month >= 8
    if r["n"] >= 4 and ativo_ago: return ("dor que aperta", "#b71c1c")
    if r["a"] and r["a"].month <= 4 and ativo_ago: return ("crônico desde o início", "#c77d1f")
    if r["n"] >= 3: return ("recorrente", "#1a4a8a")
    return ("sinal isolado", "#5b626a")
top = sorted([r for r in rows if r["a"]], key=lambda r: (-r["n"], r["a"]))[:8]
MN = ["", "jan", "fev", "mar", "abr", "mai", "jun", "jul", "ago", "set", "out", "nov", "dez"]
cards = ""
for r in top:
    lab, colr = badge(r)
    m = re.search(r"F(\d+):\s*['\u2018\u2019\"]([^'\u2018\u2019\"]{12,220})['\u2018\u2019\"]", str(r["cit"] or ""))
    quote = ('<p class="q">"%s"<sup>%d</sup></p>' % (esc(m.group(2).strip()), int(m.group(1)))) if m else ""
    sups = ", ".join(str(int(f[1:])) for f in re.findall(r"F\d+", r["fontes"]))
    como = esc(r["como"] or "").strip()
    como = como if len(como) < 420 else como[:417].rsplit(" ", 1)[0] + "."
    cards += '<div class="news"><p class="nt"><strong>%s. %s</strong> <span class="badge" style="background:%s">%s</span><span class="meta">%d documento%s · %s a %s · %s</span></p><p>%s<sup>%s</sup></p>%s</div>\n' % (
        esc(r["id"]), esc(r["sint"]), colr, lab, r["n"], "s" if r["n"] != 1 else "", MN[r["a"].month], MN[(r["b"] or r["a"]).month], esc(r["ramo"]), como, sups, quote)
SEC_C = '''
<h3>O que estamos cobrindo: os sintomas que mais aparecem, um a um</h3>
<p>Os oito sintomas citados em mais documentos, cada um com a natureza (selo colorido), os números (documentos que o citam, período em que esteve ativo, ramo afetado), como se manifesta e uma fala da ata.</p>
''' + cards + '''<p>Falas conferidas nas atas. ''' + SV + ''' A natureza de cada sintoma é leitura deste documento. ''' + SI + '''</p>
<p class="note">O que as atas mostram e nenhuma reunião diz em voz alta está na aba Riscos, em "O que ninguém está falando".</p>
'''
B = load("frag_B.html")
anchor = re.search(r'<section id="p-sint" class="pane">.*?(?:<figure style="margin:14px 0 22px">.*?</figure>\n|<p class="gancho">.*?</p>\n)', B, re.S)
if anchor: B = B.replace(anchor.group(0), anchor.group(0) + FIG_A + FIG_B + SEC_C, 1); print("sint ok; cartões:", len(top))
else: print("MISS sint")
save("frag_B.html", B)
b = load("build.py")
if ".news{" not in b:
    b = b.replace(".sel-p{background:#e8f0fe;color:#1a4a8a}", ".sel-p{background:#e8f0fe;color:#1a4a8a} .news{border:1px solid #e3e3e3;border-radius:8px;padding:12px 18px;margin:10px 0;background:#fff} .news .nt{margin:0 0 6px;font-size:15.5px} .news p{margin:6px 0} .badge{display:inline-block;color:#fff;font-family:system-ui,sans-serif;font-size:11px;font-weight:700;padding:2px 8px;border-radius:10px;margin-left:8px;vertical-align:2px} .meta{float:right;font-family:system-ui,sans-serif;font-size:11.5px;color:#777} .news .q{border-left:3px solid #ddd;padding-left:12px;font-style:italic;color:#555;font-size:13.5px}", 1)
save("build.py", b)
print("patch v26 ok")

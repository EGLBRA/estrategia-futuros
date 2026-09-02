# -*- coding: utf-8 -*-
"""v29: gráfico problema x quantidade de documentos em que aparece, com recorte organizacional (cor por categoria)."""
import io, os, re
import openpyxl
W = r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01"
X = r"A:\_01 Projetos\Estrategia\Pesquisas\Sintomas Organizacionais GT Susep Vida e RE - 26-09-01\Sintomas_Organizacionais_Susep_Vida_RE.xlsx"
def load(n): return io.open(os.path.join(W, n), encoding="utf-8").read()
def save(n, s): io.open(os.path.join(W, n), "w", encoding="utf-8", newline="\n").write(s)
SV = '<span class="sel sel-v">Verificado</span>'
def esc(t): return str(t).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
wb = openpyxl.load_workbook(X, data_only=True); ws = wb["Sintomas"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ic, isn, inn = hdr.index("Categoria"), hdr.index("Sintoma / Problema"), hdr.index("Nº de fontes")
rows = [(r[isn], r[ic], int(r[inn] or 0)) for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
CC = {"Organização do Trabalho": "#1a4a8a", "Governança": "#b71c1c", "Coordenação": "#b47c00", "Informação": "#6410ab", "Recursos": "#00995d", "Tecnologia": "#5b626a"}
top = sorted(rows, key=lambda r: -r[2])[:16]
mx = max(r[2] for r in top)
H = 60 + len(top) * 30 + 70
s = '<svg viewBox="0 0 1100 %d" role="img" aria-label="Os dezesseis problemas que mais aparecem nas atas, com a cor indicando a dimensão organizacional" xmlns="http://www.w3.org/2000/svg" style="display:block;width:100%%;height:auto;font-family:system-ui,sans-serif">' % H
s += '<text x="20" y="28" font-size="15" font-weight="700" fill="#1a1a1a">Problema e quantidade de documentos em que aparece</text><text x="20" y="48" font-size="12.5" fill="#5a6068">os dezesseis mais citados entre os 72; a cor é o recorte organizacional</text>'
x0 = 470; bw = 560
for i, (name, cat, n) in enumerate(top):
    y = 64 + i * 30
    lab = esc(name if len(str(name)) <= 62 else str(name)[:59].rsplit(" ", 1)[0] + "...")
    s += '<text x="%d" y="%d" text-anchor="end" font-size="12.5" fill="#1a1a1a">%s</text>' % (x0 - 10, y + 15, lab)
    w = bw * n / mx
    s += '<rect x="%d" y="%d" width="%d" height="20" rx="3" fill="%s"/>' % (x0, y + 2, w, CC.get(cat, "#999"))
    s += '<text x="%d" y="%d" font-size="12.5" font-weight="700" fill="#1a1a1a">%d</text>' % (x0 + w + 8, y + 16, n)
ly = 64 + len(top) * 30 + 22
lx = 20
for cat, c in CC.items():
    s += '<rect x="%d" y="%d" width="14" height="14" rx="3" fill="%s"/><text x="%d" y="%d" font-size="12" fill="#1a1a1a">%s</text>' % (lx, ly, c, lx + 20, ly + 12, cat)
    lx += 24 + 8 * len(cat) + 20
s += '</svg>'
FIG = '<figure style="margin:14px 0 22px">\n' + s + '\n<figcaption class="note" style="margin-top:6px">Contagem de documentos que citam cada problema, na planilha de sintomas.<sup>25</sup> Os mais citados se espalham pelas seis dimensões: o padrão é do desenho, não de uma área. ' + SV + '</figcaption>\n</figure>\n'
B = load("frag_B.html")
m = re.search(r'<h3>O que estamos cobrindo: os sintomas que mais aparecem, um a um</h3>', B)
if m: B = B.replace(m.group(0), FIG + m.group(0), 1); print("ok", len(top), mx)
else: print("MISS")
save("frag_B.html", B)

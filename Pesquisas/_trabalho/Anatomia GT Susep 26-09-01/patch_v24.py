# -*- coding: utf-8 -*-
"""v24: gráfico de temperatura dos sintomas (categoria x mês, da planilha); Loops vira Laços; apex sem 'para o leigo';
gancho da manchete sem tom agressivo."""
import io, os, re, datetime
import openpyxl
W = r"A:\_01 Projetos\Estrategia\Pesquisas\_trabalho\Anatomia GT Susep 26-09-01"
X = r"A:\_01 Projetos\Estrategia\Pesquisas\Sintomas Organizacionais GT Susep Vida e RE - 26-09-01\Sintomas_Organizacionais_Susep_Vida_RE.xlsx"
def load(n): return io.open(os.path.join(W, n), encoding="utf-8").read()
def save(n, s): io.open(os.path.join(W, n), "w", encoding="utf-8", newline="\n").write(s)
SI = '<span class="sel sel-i">Inferência</span>'; SV = '<span class="sel sel-v">Verificado</span>'
MISS = []

# ---- 1. temperatura: sintomas ativos por categoria e mês (entre primeira e última menção)
wb = openpyxl.load_workbook(X, data_only=True); ws = wb["Sintomas"]
hdr = [str(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
ic, i1, i2 = hdr.index("Categoria"), hdr.index("Primeira menção"), hdr.index("Última menção")
def d(v):
    if isinstance(v, datetime.datetime): return v.date()
    if isinstance(v, datetime.date): return v
    try: return datetime.datetime.strptime(str(v)[:10], "%d/%m/%Y").date()
    except Exception: return None
months = [(2026, m) for m in range(3, 9)]; mname = ["mar", "abr", "mai", "jun", "jul", "ago"]
cats = ["Organização do Trabalho", "Governança", "Coordenação", "Informação", "Recursos", "Tecnologia"]
grid = {c: [0] * 6 for c in cats}; n = 0
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[0] or r[ic] not in grid: continue
    a, b = d(r[i1]), d(r[i2])
    if not a: continue
    b = b or a; n += 1
    for k, (y, m) in enumerate(months):
        first = datetime.date(y, m, 1); last = datetime.date(y, m + 1, 1) - datetime.timedelta(days=1)
        if a <= last and b >= first: grid[r[ic]][k] += 1
mx = max(max(v) for v in grid.values())
def col(v):
    if v == 0: return "#f4f5f7"
    t = v / mx; r_, g_, b_ = 255, int(236 - 170 * t), int(230 - 190 * t)
    return "#%02x%02x%02x" % (r_, g_, b_)
s = '<svg viewBox="0 0 1100 330" role="img" aria-label="Temperatura dos sintomas: quantos sintomas de cada categoria estavam ativos em cada mês, de março a agosto de 2026" xmlns="http://www.w3.org/2000/svg" style="display:block;width:100%;height:auto;font-family:system-ui,sans-serif">'
s += '<text x="20" y="28" font-size="15" font-weight="700" fill="#1a1a1a">Temperatura dos sintomas, março a agosto de 2026</text><text x="20" y="48" font-size="12.5" fill="#5a6068">cada célula conta os sintomas da categoria ativos no mês (entre a primeira e a última menção nas atas); quanto mais escuro, mais quente</text>'
x0, y0, cw, ch = 300, 70, 120, 36
for k, mn in enumerate(mname):
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="13" font-weight="600" fill="#5a6068">%s</text>' % (x0 + k * cw + cw / 2, y0 - 8, mn)
tot = [0] * 6
for i, c in enumerate(cats):
    y = y0 + i * ch
    s += '<text x="20" y="%d" font-size="13.5" fill="#1a1a1a">%s</text>' % (y + 23, c)
    for k, v in enumerate(grid[c]):
        tot[k] += v
        s += '<rect x="%d" y="%d" width="%d" height="%d" fill="%s" stroke="#fff"/>' % (x0 + k * cw, y, cw, ch, col(v))
        s += '<text x="%d" y="%d" text-anchor="middle" font-size="13.5" font-weight="600" fill="%s">%d</text>' % (x0 + k * cw + cw / 2, y + 23, "#1a1a1a" if v < mx * 0.6 else "#fff", v)
y = y0 + 6 * ch + 6
s += '<line x1="20" y1="%d" x2="1040" y2="%d" stroke="#d7dbe0"/>' % (y, y)
s += '<text x="20" y="%d" font-size="13.5" font-weight="700" fill="#1a1a1a">Sintomas ativos no mês</text>' % (y + 24)
for k, v in enumerate(tot):
    s += '<text x="%d" y="%d" text-anchor="middle" font-size="14" font-weight="700" fill="#b71c1c">%d</text>' % (x0 + k * cw + cw / 2, y + 24, v)
s += '</svg>'
FIG = '<figure style="margin:14px 0 22px">\n' + s + '\n<figcaption class="note" style="margin-top:6px">Contagem deste documento a partir das datas de primeira e última menção de cada um dos %d sintomas na planilha.<sup>25</sup> A temperatura sobe de março a agosto: o número de sintomas ativos cresce e não cai. ' % n + SI + '</figcaption>\n</figure>\n'
B = load("frag_B.html")
m = re.search(r'<section id="p-sint" class="pane">.*?<p class="gancho">.*?</p>\n', B, re.S)
if m: B = B.replace(m.group(0), m.group(0) + FIG, 1)
else: MISS.append("sint gancho")
save("frag_B.html", B)

# ---- 2. Loops vira Laços
b = load("build.py"); b = b.replace('("loops", "Loops")', '("loops", "Laços")'); save("build.py", b)
for fn in [f"frag_{k}.html" for k in "ABCDEFGHIJKL"] + ["pratica.py"]:
    try: t = load(fn)
    except FileNotFoundError: continue
    t = t.replace("<h2>Loops: três laços que explicam por que o problema volta</h2>", "<h2>Laços: por que o problema volta</h2>").replace("aba Loops", "aba Laços").replace("abas Loops", "abas Laços")
    save(fn, t)

# ---- 3. apex sem 'para o leigo'
A = load("frag_A.html")
if "<strong>Para o leigo, em uma imagem.</strong>" in A: A = A.replace("<strong>Para o leigo, em uma imagem.</strong>", "<strong>Em uma imagem.</strong>")
else: MISS.append("apex")
A = A.replace("O que muda entre uma esteira e outra, se não são as pessoas?", "O que muda entre uma esteira e outra?")
save("frag_A.html", A)

# ---- 4. manchete sem tom agressivo
b = load("build.py")
old = "Por que as mesmas pessoas que fazem Ramos Elementares crescer não conseguem destravar Vida?"
if old in b: b = b.replace(old, "O que faz Ramos Elementares avançar, e o que falta para Vida avançar do mesmo jeito?")
else: MISS.append("gancho manchete")
save("build.py", b)
print("patch v24 ok; faltas:", MISS, "| sintomas com data:", n, "| max:", mx)
